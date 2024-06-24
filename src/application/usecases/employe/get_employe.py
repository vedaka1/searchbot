from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook, load_workbook

from application.common.workbook import EmployeWorkbook
from domain.common.response import Response


@dataclass
class GetEmploye:
    wb: EmployeWorkbook
    _ignored_symbols: list = field(
        default=[None, "", "@", "-", "9999999"],
        init=False,
    )

    def __call__(self, search_prompt: str) -> str:

        ws = self.wb.active
        result = ""
        search_columns = ws["N":"Q"]
        try:
            for cell in tuple(zip(search_columns[0], search_columns[1])):
                if (
                    search_prompt.lower() in str(cell[0].value).lower()
                    or search_prompt.lower() in str(cell[1].value).lower()
                ):
                    result += "\n"
                    row = ws[cell[0].row]
                    cells_to_parse = (2, 4, 7)
                    for cell_number in cells_to_parse:
                        if row[cell_number].value not in self._ignored_symbols:
                            result += f"{row[cell_number].value}\n"

                    position = "*Должность*: {0}".format(row[16].value)
                    fullname = "*ФИО*: {0} {1}".format(row[13].value, row[14].value)
                    number = " *Номер*: ({0}){1}".format(row[18].value, row[19].value)
                    result += "{0}\n{1}\n{2}".format(position, fullname, number)

        except Exception as e:
            print(e)
            return "Записей не найдено"

        result = Response(result).value
        if not result:
            return "Записей не найдено"

        if len(result) > 4000:
            return result[:4000] + "\.\.\."
        return result
